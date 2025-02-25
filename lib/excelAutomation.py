import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
        
class ExcelAutomation:
    def __init__(self):
        pass

    @classmethod
    def replicate_merges(src_ws, dest_ws,
                        src_start_col, src_start_row,
                        src_end_col,   src_end_row,
                        dest_start_col, dest_start_row):


        start_col_idx = column_index_from_string(src_start_col)
        end_col_idx   = column_index_from_string(src_end_col)
        dest_start_col_idx = column_index_from_string(dest_start_col)

        # Offsets from source range to destination range
        row_offset = dest_start_row - src_start_row
        col_offset = dest_start_col_idx - start_col_idx

        # Each item in src_ws.merged_cells.ranges is something like "B1:D2"
        for merge_range in src_ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))

            # Check if this merged range intersects our copy range
            intersects_horizontally = not (max_col < start_col_idx or min_col > end_col_idx)
            intersects_vertically   = not (max_row < src_start_row or min_row > src_end_row)

            if intersects_horizontally and intersects_vertically:
                # Calculate corresponding new range in destination
                new_min_col = min_col + col_offset
                new_max_col = max_col + col_offset
                new_min_row = min_row + row_offset
                new_max_row = max_row + row_offset

                # Merge the corresponding cells in dest_ws
                tl = f"{get_column_letter(new_min_col)}{new_min_row}"  # top-left
                br = f"{get_column_letter(new_max_col)}{new_max_row}"  # bottom-right
                dest_ws.merge_cells(f"{tl}:{br}")

    @classmethod
    def copy_block(src_ws, dest_ws,
                src_start_col, src_start_row,
                src_end_col,   src_end_row,
                dest_start_col, dest_start_row):
        
        ExcelAutomation.replicate_merges(src_ws, dest_ws,
                        src_start_col, src_start_row,
                        src_end_col,   src_end_row,
                        dest_start_col, dest_start_row)
        
        start_col_index = column_index_from_string(src_start_col)
        end_col_index = column_index_from_string(src_end_col)
        dest_start_col_index = column_index_from_string(dest_start_col)

        row_count = src_end_row - src_start_row + 1
        col_count = end_col_index - start_col_index + 1

        for row_offset in range(row_count):
            for col_offset in range(col_count):
                src_cell = src_ws.cell(
                    row = src_start_row + row_offset,
                    column = start_col_index + col_offset
                )
                dest_cell = dest_ws.cell(
                    row = dest_start_row + row_offset,
                    column = dest_start_col_index + col_offset
                )

                if isinstance(dest_cell, MergedCell):
                    continue
                
                dest_cell.value = src_cell.value
                dest_cell.alignment = Alignment(horizontal='center', vertical='center')


    @classmethod
    def set_font(sheet, min_row, max_row, min_col, max_col, size, bold=False):
        """Helper to set font size & bold in a cell range."""
        for row_cells in sheet.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col
        ):
            for cell in row_cells:
                cell.font = Font(size=size, bold=bold)

    @classmethod
    def fill_x_if_empty(sheet, row, start_col, end_col):
        """
        For the specified sheet, row, and column range, 
        write 'X' if the cell value is None (empty).
        """
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is None:
                cell.value = "X"

    @classmethod
    def automate(source_file, source_sheet, output_file):

        # Load the workbooks
        source_wb = openpyxl.load_workbook(source_file)
        source_ws = source_wb[source_sheet]

        # Open or create the output workbook
        output_wb = openpyxl.load_workbook(output_file)

        max_row = source_ws.max_row
        sheet_index = 0

        for row_idx in range(1, max_row + 1):
            name_value = source_ws.cell(row=row_idx, column=1).value

            if not name_value:
                continue

            if row_idx < 5:
                continue

            if sheet_index == 0:
                dest_sheet = output_wb.worksheets[0]
                dest_sheet.title = str(name_value)
            else:
                if sheet_index < len(output_wb.worksheets):
                    dest_sheet = output_wb.worksheets[sheet_index]
                    dest_sheet.title = str(name_value)
                else:
                    dest_sheet = output_wb.create_sheet(title=str(name_value))
            
            # Writing header
            dest_sheet["A1"] = f"{name_value} 특강 학습 확인표"
            dest_sheet.merge_cells("A1:H1")
            dest_sheet["A1"].font = Font(size=18, bold=True)
            dest_sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')

            src_block_top = row_idx - 4
            src_block_bottom = row_idx

            ExcelAutomation.copy_block(source_ws, dest_sheet, "B", src_block_top, "I", src_block_bottom, "A", 2)
            ExcelAutomation.copy_block(source_ws, dest_sheet, "J", src_block_top, "Q", src_block_bottom, "A", 7)
            ExcelAutomation.copy_block(source_ws, dest_sheet, "R", src_block_top, "Y", src_block_bottom, "A", 12)

            ExcelAutomation.set_font(dest_sheet, 2,  4,   1, 8, 11, False)
            ExcelAutomation.set_font(dest_sheet, 7,  9,   1, 8, 11, False)
            ExcelAutomation.set_font(dest_sheet, 12, 14,  1, 8, 11, False)
            ExcelAutomation.set_font(dest_sheet, 6,  6,   1, 8, 11, True )
            ExcelAutomation.set_font(dest_sheet, 11, 11,  1, 8, 11, True )
            ExcelAutomation.set_font(dest_sheet, 16, 16,  1, 8, 11, True )

            ExcelAutomation.fill_x_if_empty(dest_sheet,  6, 1, 8)
            ExcelAutomation.fill_x_if_empty(dest_sheet, 11, 1, 8)
            ExcelAutomation.fill_x_if_empty(dest_sheet, 16, 1, 8)

            sheet_index += 1

        output_wb.save(output_file)
        print(f"Automation complete, saved to {output_file}")
